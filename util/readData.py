import openpyxl as xlsx
import numpy as np
import json
import os
from .path import *

workbook = xlsx.load_workbook(filename=FEED_DATA_PATH)
sheet = workbook.get_sheet_by_name(name='Big Data Structured Matrix')
business_areas = open(BUSINESS_AREAS_PATH, 'r')
suppliers_data_columns = "EFGHIJK"

def getCell(column, row):
    """
    """
    cellName = '{}{}'.format(column, row)
    cellValue = sheet[cellName].value
    return cellValue


def getSupplierCode(row):
    """
    """
    return getCell('B', row)


def getBusinessArea(row):
    """
    """
    return getCell('D', row)


def getBusinessAreasList():
    """
    """
    return json.loads(business_areas.read())


def getColumn(column):
    """
    (String) -> (List)

    Gets the column from Big Data Structured Matrix sheet with the name of '@param:column'
    """
    result = []
    for row in range(3, sheet.max_row):
        result.append(getCell(column, row))

    return result


def getRow(row):
    """
    """
    data = []
    for column in suppliers_data_columns:
        data.append(getCell(column, row))

    return np.array(data)


def getSuppliersData(businessArea):
    """
    """
    suppliers = {}
    for row in range(3, sheet.max_row):
        if not getBusinessArea(row) == businessArea:
            continue
        hasNullData = False
        sDataList = []
        for column in suppliers_data_columns:
            cellName = "{}{}".format(column, row)
            cellValue = sheet[cellName].value
            if cellValue is None:
                hasNullData = True
                break
            sDataList.append(cellValue)

        if hasNullData:
            continue

        sDataArray = np.array(sDataList)
        suppliers[row] = sDataArray

    return suppliers